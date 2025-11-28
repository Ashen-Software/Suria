"""
Parser especializado para Excel de Declaración de Producción de Gas Natural - MinMinas.

Este módulo parsea la estructura compleja del Excel:
- Filas de cabecera (campo, contrato, operador, poder calorífico, períodos)
- Datos mensuales por operador y tipo de producción (120 columnas = 10 años × 12 meses)
- Filas de participación del Estado

Estructura del Excel (basado en análisis de res_00014.csv):
- Fila 5: "DECLARACIÓN DE PRODUCCIÓN", "OPERADOR DEL CAMPO", "DESDE", ...
- Fila 6: vacío, {operador_campo}, "HASTA", ...
- Fila 7: "CAMPO", "PODER CALORIFICO (BTU/PC)", "CONTRATO"
- Fila 8: {nombre_campo}, {poder_calorifico}, {contrato}
- Fila 9: Headers de años: "2021 - Año 1 (GBTUD)", "2022 - Año 2 (GBTUD)", ...
- Fila 10: Headers de meses: "Ene", "Feb", ..., "Dic" × 10 años
- Filas 11+: Datos por operador y tipo de producción
"""
from typing import Dict, Any, List, Optional, Tuple
from datetime import date
from decimal import Decimal, InvalidOperation
from dataclasses import dataclass
import re
import openpyxl
from io import BytesIO

from logs_config.logger import app_logger as logger


@dataclass
class CampoMetadata:
    """Metadatos del campo extraídos de las cabeceras del Excel."""
    nombre_campo: str
    contrato: str
    operador_campo: str
    poder_calorifico_btu_pc: Optional[float]
    periodo_desde: date
    periodo_hasta: date


@dataclass
class ProduccionRecord:
    """Registro individual de producción mensual."""
    campo_nombre: str
    operador: str
    tipo_produccion: str
    es_participacion_estado: bool
    anio: int
    mes: int
    valor_gbtud: float
    poder_calorifico_btu_pc: Optional[float]
    # Metadata de resolución (viene del JSON, no del Excel)
    resolucion_number: Optional[str] = None


# Mapeo de tipos de producción del Excel a códigos de base de datos
TIPO_PRODUCCION_MAP = {
    "PTDV": "PTDV",
    "PC- Contratos suministro consumo interno": "PC_CONTRATOS",
    "PC- Exportaciones": "PC_EXPORTACIONES",
    "PC- Refinería de Barrancabermeja": "PC_REF_BARRANCA",
    "PC- Refinería de Cartagena": "PC_REF_CARTAGENA",
    "PP (DECLARADO POR EL OPERADOR DEL CAMPO)": "PP",
    "GAS OPERACIÓN": "GAS_OPERACION",
    # CIDV tiene formato especial: "CIDV - {operador}"
}

# Mapeo de meses español a número
MESES_MAP = {
    "Ene": 1, "Feb": 2, "Mar": 3, "Abr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Ago": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dic": 12
}


class MinMinasExcelParser:
    """
    Parser para Excel de declaración de producción de gas natural.
    
    Cada Excel representa un campo y contiene:
    - Datos mensuales de producción por operador
    - Múltiples tipos de producción (PTDV, PC, PP, etc.)
    - Participación del Estado cuando aplica
    """
    
    def __init__(self, excel_bytes: bytes, resolution_number: Optional[str] = None):
        """
        Inicializa el parser con el contenido del Excel.
        
        Args:
            excel_bytes: Contenido binario del archivo Excel
            resolution_number: Número de resolución (para trazabilidad)
        """
        self.excel_bytes = excel_bytes
        self.resolution_number = resolution_number
        self.workbook = None
        self.metadata: Optional[CampoMetadata] = None
        self.records: List[ProduccionRecord] = []
        self.errors: List[Dict[str, Any]] = []
        
        # Estructura de columnas (se determina dinámicamente)
        self.year_columns: Dict[int, Tuple[int, int]] = {}  # año -> (col_inicio, col_fin)
        self.data_start_row: int = 11  # Fila donde empiezan los datos
        
    def parse(self) -> Tuple[List[ProduccionRecord], List[Dict[str, Any]]]:
        """Parsea el Excel completo."""
        try:
            # Cargar workbook con read_only para eficiencia
            self.workbook = openpyxl.load_workbook(
                BytesIO(self.excel_bytes), 
                data_only=True,
                read_only=True
            )
            
            for sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                # Cargar todas las filas en memoria de una vez (más eficiente)
                rows = list(sheet.iter_rows(values_only=True))
                self._parse_sheet_rows(sheet.title, rows)
            
            logger.info(
                f"[MinMinasParser] Completado: {len(self.workbook.sheetnames)} hojas, "
                f"{len(self.records)} registros, {len(self.errors)} errores"
            )
            
            return self.records, self.errors
            
        except Exception as e:
            logger.error(f"[MinMinasParser] Error crítico parseando Excel: {e}")
            self.errors.append({
                "type": "critical",
                "message": str(e),
                "resolution": self.resolution_number
            })
            return [], self.errors
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _parse_sheet_rows(self, sheet_title: str, rows: List[tuple]) -> None:
        """Parsea una hoja usando lista de filas pre-cargadas."""
        try:
            if len(rows) < 11:
                self.errors.append({
                    "type": "structure",
                    "message": "Hoja muy corta, menos de 11 filas",
                    "sheet": sheet_title
                })
                return
            
            # 1. Detectar estructura de columnas (fila 9, índice 8)
            self._detect_column_structure_from_row(rows[8] if len(rows) > 8 else ())
            
            if not self.year_columns:
                self.errors.append({
                    "type": "structure",
                    "message": "No se detectaron columnas de años/meses",
                    "sheet": sheet_title
                })
                return
            
            # 2. Extraer metadatos
            self.metadata = self._extract_metadata_from_rows(sheet_title, rows)
            if not self.metadata:
                self.errors.append({
                    "type": "metadata",
                    "message": "No se pudieron extraer metadatos del campo",
                    "sheet": sheet_title
                })
                return
            
            # 3. Parsear filas de datos (desde fila 11, índice 10)
            self._parse_data_from_rows(rows[10:])
            
        except Exception as e:
            logger.error(f"[MinMinasParser] Error parseando hoja {sheet_title}: {e}")
            self.errors.append({
                "type": "sheet_error",
                "message": str(e),
                "sheet": sheet_title
            })
    
    def _detect_column_structure_from_row(self, row: tuple) -> None:
        """Detecta columnas de años desde la fila 9."""
        self.year_columns = {}
        
        for col_idx, cell_value in enumerate(row):
            if not cell_value:
                continue
            
            cell_str = str(cell_value)
            match = re.search(r'(\d{4})\s*-?\s*Año\s*\d+', cell_str)
            if match:
                year = int(match.group(1))
                # col_idx es 0-based, guardamos como 0-based
                self.year_columns[year] = (col_idx, col_idx + 11)
    
    def _extract_metadata_from_rows(self, sheet_title: str, rows: List[tuple]) -> Optional[CampoMetadata]:
        """Extrae metadatos desde las filas pre-cargadas."""
        try:
            nombre_campo = self._extract_campo_from_sheet_title(sheet_title)
            
            # Fila 8 (índice 7): campo en col C (2), poder calorífico en col F (5)
            if len(rows) > 7:
                row8 = rows[7]
                if not nombre_campo and len(row8) > 2:
                    nombre_campo = row8[2]  # Col C
                poder_calorifico_raw = row8[5] if len(row8) > 5 else None
            else:
                poder_calorifico_raw = None
            
            poder_calorifico = self._parse_numeric(poder_calorifico_raw)
            
            if self.year_columns:
                años = sorted(self.year_columns.keys())
                periodo_desde = date(años[0], 1, 1)
                periodo_hasta = date(años[-1], 12, 31)
            else:
                periodo_desde = date(2021, 1, 1)
                periodo_hasta = date(2030, 12, 31)
            
            if not nombre_campo:
                return None
            
            return CampoMetadata(
                nombre_campo=str(nombre_campo).strip(),
                contrato="",
                operador_campo="",
                poder_calorifico_btu_pc=poder_calorifico,
                periodo_desde=periodo_desde,
                periodo_hasta=periodo_hasta
            )
        except Exception as e:
            logger.error(f"[MinMinasParser] Error extrayendo metadata: {e}")
            return None
    
    def _parse_data_from_rows(self, data_rows: List[tuple]) -> None:
        """Parsea filas de datos desde lista pre-cargada."""
        current_operador = ""
        empty_count = 0
        max_empty = 5
        
        for row in data_rows:
            if not row or len(row) < 4:
                empty_count += 1
                if empty_count >= max_empty:
                    break
                continue
            
            col_c = row[2] if len(row) > 2 else None  # Operador
            col_d = row[3] if len(row) > 3 else None  # Tipo
            
            if not col_c and not col_d:
                empty_count += 1
                if empty_count >= max_empty:
                    break
                continue
            
            empty_count = 0
            
            if col_c:
                current_operador = str(col_c).strip()
            
            tipo_str = str(col_d).strip() if col_d else ""
            
            es_estado = "ESTADO" in current_operador.upper()
            operador_str = current_operador
            if es_estado:
                match = re.match(r'ESTADO\s*-\s*(.+)', operador_str, re.IGNORECASE)
                if match:
                    operador_str = match.group(1).strip()
            
            tipo_produccion = self._map_tipo_produccion(tipo_str, operador_str)
            
            if tipo_produccion and current_operador:
                self._parse_monthly_values_from_row(row, operador_str, tipo_produccion, es_estado)
    
    def _parse_monthly_values_from_row(
        self, 
        row: tuple, 
        operador: str, 
        tipo_produccion: str, 
        es_estado: bool
    ) -> None:
        """Parsea valores mensuales desde una fila."""
        for year, (col_start, col_end) in sorted(self.year_columns.items()):
            for month_offset in range(12):
                col = col_start + month_offset
                if col > col_end or col >= len(row):
                    break
                
                valor = self._parse_numeric(row[col])
                if valor is None:
                    valor = 0.0
                
                if valor > 0:
                    self.records.append(ProduccionRecord(
                        campo_nombre=self.metadata.nombre_campo,
                        operador=operador,
                        tipo_produccion=tipo_produccion,
                        es_participacion_estado=es_estado,
                        anio=year,
                        mes=month_offset + 1,
                        valor_gbtud=valor,
                        poder_calorifico_btu_pc=self.metadata.poder_calorifico_btu_pc,
                        resolucion_number=self.resolution_number
                    ))
    
    def _map_tipo_produccion(self, tipo_str: str, operador_str: str) -> Optional[str]:
        """
        Mapea el tipo de producción del Excel al código de base de datos.
        
        Args:
            tipo_str: Texto del tipo de producción en el Excel
            operador_str: Operador (para detectar CIDV - {operador})
            
        Returns:
            Código de tipo de producción o None si no se reconoce
        """
        if not tipo_str:
            return None
        
        tipo_upper = tipo_str.upper().strip()
        
        # Mapeos directos
        if tipo_upper == "PTDV":
            return "PTDV"
        if "CONTRATOS" in tipo_upper and "CONSUMO" in tipo_upper:
            return "PC_CONTRATOS"
        if "EXPORTACIONES" in tipo_upper:
            return "PC_EXPORTACIONES"
        if "BARRANCABERMEJA" in tipo_upper or "BARRANCA" in tipo_upper:
            return "PC_REF_BARRANCA"
        if "CARTAGENA" in tipo_upper:
            return "PC_REF_CARTAGENA"
        if tipo_upper.startswith("PP") or "POTENCIAL" in tipo_upper:
            return "PP"
        if "GAS OPERACIÓN" in tipo_upper or "GAS OPERACION" in tipo_upper:
            return "GAS_OPERACION"
        if tipo_upper.startswith("CIDV"):
            return "CIDV"
        
        # No reconocido
        return None
    
    def _get_cell_value(self, sheet, row: int, col: int) -> Any:
        """Obtiene el valor de una celda de forma segura."""
        try:
            cell = sheet.cell(row=row, column=col)
            return cell.value
        except Exception:
            return None
    
    def _parse_numeric(self, value: Any) -> Optional[float]:
        """Parsea un valor numérico de forma segura."""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        try:
            # Limpiar string: remover comas de miles, espacios
            value_str = str(value).strip()
            value_str = value_str.replace(",", "")  # Formato europeo usa . como miles
            value_str = value_str.replace(" ", "")
            
            if not value_str or value_str == "-":
                return None
            
            return float(value_str)
        except (ValueError, InvalidOperation):
            return None
    
    def _extract_campo_from_sheet_title(self, title: str) -> Optional[str]:
        """
        Extrae el nombre del campo del título de la hoja.
        
        Formatos esperados:
        - "VIGIA SUR - 1" -> "VIGIA SUR"
        - "CAMPO RICO - NORTE - 2" -> "CAMPO RICO"
        - "LA CRECIENTE" -> "LA CRECIENTE"
        
        El nombre es el primer segmento antes del "-".
        """
        if not title:
            return None
        
        title = str(title).strip()
        
        # Si tiene "-", tomar el primer segmento
        if "-" in title:
            nombre = title.split("-")[0].strip()
            if nombre:
                return nombre
        
        # Si no tiene "-", usar el título completo
        return title
    
    def _parse_date(self, date_str: str) -> Optional[date]:
        """Parsea una fecha en varios formatos posibles."""
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        
        # Formatos comunes
        formats = [
            "%d/%m/%Y",   # 01/01/2021
            "%Y-%m-%d",   # 2021-01-01
            "%d-%m-%Y",   # 01-01-2021
            "%Y/%m/%d",   # 2021/01/01
        ]
        
        from datetime import datetime
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        # Intentar extraer año si todo falla
        match = re.search(r'(\d{4})', date_str)
        if match:
            year = int(match.group(1))
            # Determinar mes/dia por contexto
            if "01/01" in date_str or "enero" in date_str.lower():
                return date(year, 1, 1)
            if "31/12" in date_str or "diciembre" in date_str.lower():
                return date(year, 12, 31)
        
        return None
