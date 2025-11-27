"""
Módulo para parseo y extracción de datos de archivos Excel de declaraciones.
"""
from typing import Dict, Any, List, Optional
from io import BytesIO
from datetime import datetime
import traceback

from openpyxl import load_workbook
from logs_config.logger import app_logger as logger
from extraction.scrapers.declaracion.utils import get_spanish_months_map


def extract_excel_data(excel_bytes: BytesIO, declaration: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Extrae datos estructurados del archivo Excel.
    
    Args:
        excel_bytes: BytesIO con el contenido del Excel
        declaration: Información de la declaración para metadata
        
    Returns:
        Lista de diccionarios con datos extraídos
    """
    extracted_records = []
    
    try:
        workbook = load_workbook(excel_bytes, read_only=True, data_only=True)
        
        logger.info(f"[excel_parser] Procesando workbook con {len(workbook.sheetnames)} hojas")
        
        for sheet_name in workbook.sheetnames:
            try:
                sheet = workbook[sheet_name]
                sheet_data = parse_excel_sheet(sheet, sheet_name, declaration)
                extracted_records.extend(sheet_data)
                
            except Exception as e:
                logger.warning(f"[excel_parser] Error procesando hoja {sheet_name}: {e}")
                continue
        
        workbook.close()
        
        logger.info(f"[excel_parser] Extraídos {len(extracted_records)} registros del Excel")
        
    except Exception as e:
        logger.error(f"[excel_parser] Error leyendo Excel: {e}")
    
    return extracted_records


def parse_excel_sheet(sheet, sheet_name: str, declaration: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Parsea una hoja específica del Excel y extrae datos estructurados.
    
    Basado en la estructura observada:
    - Campo (campo)
    - Contrato (contrato)
    - Operador (operador del campo)
    - Poder calorífico (BTU/PC)
    - Datos mensuales por entidad y categoría
    
    Args:
        sheet: Hoja de openpyxl
        declaration: Metadata de la declaración
        
    Returns:
        Lista de registros extraídos
    """
    records = []
    
    try:
        metadata = extract_sheet_metadata(sheet)
        
        if not metadata.get("field_name") or not metadata.get("contract"):
            logger.warning(f"[excel_parser] No se encontró campo o contrato en hoja {sheet_name}")
            return records
        
        data_rows = extract_data_rows(sheet)
        
        for data_row in data_rows:
            record = {
                "sheet_name": sheet_name,
                "field": metadata.get("field_name"),
                "contract": metadata.get("contract"),
                "operator": metadata.get("operator"),
                "calorific_power_btu_pc": metadata.get("calorific_power"),
                "period": declaration.get("period"),
                "resolution_number": (
                    declaration.get("resolution", {}).get("number")
                    if declaration.get("resolution") else None
                ),
                "resolution_date": (
                    declaration.get("resolution", {}).get("date")
                    if declaration.get("resolution") else None
                ),
                "entity": data_row.get("entity"),
                "category": data_row.get("category"),
                "monthly_data": data_row.get("monthly_data", {}),
                "extraction_timestamp": datetime.now().isoformat()
            }
            records.append(record)
        
    except Exception as e:
        logger.error(f"[excel_parser] Error parseando hoja {sheet_name}: {e}")
    
    return records


def extract_sheet_metadata(sheet) -> Dict[str, Optional[str]]:
    """
    Extrae metadatos de la hoja Excel (campo, contrato, operador, poder calorífico).
    
    Args:
        sheet: Hoja de openpyxl
        
    Returns:
        Diccionario con metadatos extraídos
    """
    metadata = {
        "field_name": None,
        "contract": None,
        "operator": None,
        "calorific_power": None,
        "period_start": None
    }
    
    try:
        for row in sheet.iter_rows(min_row=1, max_row=15, values_only=False):
            row_values = [str(cell.value).lower() if cell.value else "" for cell in row]
            row_str = " ".join(row_values)
            
            if not row_str.strip():
                continue
            
            for cell in row:
                if not cell.value:
                    continue
                
                cell_str = str(cell.value).lower().strip()
                
                if "campo" in cell_str and not metadata["field_name"]:
                    metadata["field_name"] = _extract_next_cell_value(sheet, cell)
                
                if "contrato" in cell_str and not metadata["contract"]:
                    metadata["contract"] = _extract_next_cell_value(sheet, cell)
                
                if "operador" in cell_str and not metadata["operator"]:
                    metadata["operator"] = _extract_next_cell_value(sheet, cell)
                
                if ("poder calorifico" in cell_str or "btu/pc" in cell_str) and not metadata["calorific_power"]:
                    value_str = _extract_next_cell_value(sheet, cell)
                    if value_str:
                        try:
                            metadata["calorific_power"] = float(
                                value_str.replace(",", ".").replace(" ", "")
                            )
                        except (ValueError, TypeError):
                            pass
                
                if ("año" in cell_str or "gbtud" in cell_str) and not metadata["period_start"]:
                    metadata["period_start"] = cell_str
                    
    except Exception as e:
        logger.warning(f"[excel_parser] Error extrayendo metadatos: {e}")
    
    return metadata


def _extract_next_cell_value(sheet, cell, max_offset: int = 3) -> Optional[str]:
    """
    Extrae el valor de una celda adyacente intentando varios offsets.
    
    Args:
        sheet: Hoja de openpyxl
        cell: Celda de referencia
        max_offset: Número máximo de filas hacia abajo para buscar
        
    Returns:
        Valor de la celda encontrada o None
    """
    for offset in range(1, max_offset + 1):
        try:
            next_cell = sheet.cell(cell.row + offset, cell.column)
            if next_cell.value and str(next_cell.value).strip():
                return str(next_cell.value).strip()
        except Exception:
            continue
    return None


def extract_data_rows(sheet) -> List[Dict[str, Any]]:
    """
    Extrae filas de datos de la hoja Excel.
    Busca entidades y categorías con datos mensuales.
    
    Estructura esperada:
    - Columna B o C: Entidad (merged cells)
    - Columna C o D: Categoría
    - Columnas E en adelante: Datos mensuales
    
    Args:
        sheet: Hoja de openpyxl
        
    Returns:
        Lista de diccionarios con datos de filas
    """
    data_rows = []
    
    try:
        months_map = get_spanish_months_map()
        
        header_row, month_start_col = _find_month_header_row(sheet)
        
        if not header_row or not month_start_col:
            logger.warning("[excel_parser] No se encontró fila de encabezado de meses")
            return data_rows
        
        months_col_map = _build_month_column_map(sheet, header_row, month_start_col)
        
        if not months_col_map:
            logger.warning("[excel_parser] No se encontraron columnas de meses")
            return data_rows
        
        current_entity = None
        entity_col = 2
        category_col = 3
        
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=False):
            if row[0].row > 500:
                break
            
            row_values = [cell.value for cell in row]
            
            if not any(v for v in row_values if v):
                continue
            
            current_entity = _update_entity_from_row(
                row, entity_col, current_entity
            )
            
            category_value = _extract_category_from_row(row, category_col)
            
            if category_value and current_entity:
                monthly_data = _extract_monthly_data(row, months_col_map)
                
                if any(v != 0.0 for v in monthly_data.values()):
                    data_rows.append({
                        "entity": current_entity,
                        "category": category_value,
                        "monthly_data": monthly_data
                    })
        
    except Exception as e:
        logger.error(f"[excel_parser] Error extrayendo filas de datos: {e}")
        traceback.print_exc()
        raise
    
    return data_rows


def _find_month_header_row(sheet):
    """
    Encuentra la fila de encabezado que contiene los meses.
    
    Returns:
        Tupla (header_row, month_start_col) o (None, None)
    """
    months_map = get_spanish_months_map()
    
    for row in sheet.iter_rows(min_row=1, max_row=15, values_only=False):
        row_values = [str(cell.value).lower().strip() if cell.value else "" for cell in row]
        row_str = " ".join(row_values)
        
        if any(month in row_str for month in months_map.keys()):
            month_start_col = None
            
            for col_idx, cell in enumerate(row, 1):
                if cell.value:
                    cell_str = str(cell.value).lower().strip()
                    if any(month in cell_str for month in months_map.keys()):
                        if month_start_col is None:
                            month_start_col = col_idx
                        break
            
            if month_start_col:
                return row[0].row, month_start_col
    
    return None, None


def _build_month_column_map(sheet, header_row: int, month_start_col: int) -> Dict[str, int]:
    """
    Construye un mapa de meses a columnas.
    
    Args:
        sheet: Hoja de openpyxl
        header_row: Número de fila del encabezado
        month_start_col: Columna donde empiezan los meses
        
    Returns:
        Diccionario {mes: columna}
    """
    months_map = get_spanish_months_map()
    months_col_map = {}
    
    header_row_data = list(sheet.iter_rows(min_row=header_row, max_row=header_row))[0]
    header_cells = [cell.value for cell in header_row_data]
    
    for col_idx, cell_value in enumerate(header_cells, 1):
        if col_idx >= month_start_col and cell_value:
            cell_str = str(cell_value).lower().strip()
            for month in months_map.keys():
                if month in cell_str:
                    months_col_map[month] = col_idx
                    break
    
    return months_col_map


def _update_entity_from_row(row, entity_col: int, current_entity: Optional[str]) -> Optional[str]:
    """
    Actualiza la entidad actual basándose en el contenido de la fila.
    
    Args:
        row: Fila de openpyxl
        entity_col: Número de columna donde está la entidad
        current_entity: Entidad actual
        
    Returns:
        Nueva entidad o la actual si no cambia
    """
    if len(row) >= entity_col:
        entity_cell = row[entity_col - 1]
        if entity_cell.value:
            entity_str = str(entity_cell.value).strip()
            if entity_str and entity_str.upper() not in ["PTDV", "PC-", "ESTADO"]:
                if len(entity_str) > 3:
                    return entity_str
    return current_entity


def _extract_category_from_row(row, category_col: int) -> Optional[str]:
    """
    Extrae la categoría de una fila.
    
    Args:
        row: Fila de openpyxl
        category_col: Número de columna donde está la categoría
        
    Returns:
        Categoría extraída o None
    """
    if len(row) >= category_col:
        category_cell = row[category_col - 1]
        if category_cell.value:
            category_str = str(category_cell.value).strip()
            if category_str and category_str.upper() not in ["PTDV"]:
                return category_str
    return None


def _extract_monthly_data(row, months_col_map: Dict[str, int]) -> Dict[str, float]:
    """
    Extrae datos mensuales de una fila.
    
    Args:
        row: Fila de openpyxl
        months_col_map: Mapa de meses a columnas
        
    Returns:
        Diccionario {mes: valor}
    """
    monthly_data = {}
    
    for month, col_idx in months_col_map.items():
        if col_idx <= len(row):
            cell = row[col_idx - 1]
            if cell.value is not None:
                try:
                    if isinstance(cell.value, (int, float)):
                        value = float(cell.value)
                    else:
                        value_str = str(cell.value).replace(",", ".").replace(" ", "")
                        value = float(value_str)
                    monthly_data[month] = value
                except (ValueError, TypeError):
                    monthly_data[month] = 0.0
            else:
                monthly_data[month] = 0.0
    
    return monthly_data

